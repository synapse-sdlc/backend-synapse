"""
Feature export service: XLSX (multi-sheet) + Markdown.
"""
from io import BytesIO
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")


def _style_headers(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or "")
            max_len = max(max_len, min(len(val), 50))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)


# ── XLSX Export ──────────────────────────────────────────────────────

def export_feature_xlsx(feature, spec, plan, tests, scaffold,
                        knowledge_entries, traceability_report) -> bytes:
    wb = openpyxl.Workbook()

    # Sheet 1: Overview
    ws = wb.active
    ws.title = "Overview"
    _build_overview(ws, feature, spec, plan, tests, traceability_report)

    # Sheet 2: Epic & Stories
    if spec:
        _build_stories(wb.create_sheet("Epic & Stories"), spec, feature)

    # Sheet 3: Tasks
    if plan:
        _build_tasks(wb.create_sheet("Tasks"), plan)

    # Sheet 4: Test Cases
    if tests:
        _build_tests(wb.create_sheet("Test Cases"), tests)

    # Sheet 5: Traceability
    if traceability_report and traceability_report.get("status") == "complete":
        _build_traceability(wb.create_sheet("Traceability"), traceability_report)

    # Sheet 6: Knowledge
    if knowledge_entries:
        _build_knowledge(wb.create_sheet("Knowledge & ADRs"), knowledge_entries)

    # Sheet 7: Scaffold
    if scaffold and scaffold.get("scaffold_files"):
        _build_scaffold(wb.create_sheet("Scaffold Files"), scaffold)

    for ws in wb:
        _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_overview(ws, feature, spec, plan, tests, trace):
    ws.append(["Field", "Value"])
    _style_headers(ws, 2)

    fname = spec.get("feature_name", "") if spec else ""
    ws.append(["Feature", fname or getattr(feature, "description", "")])
    ws.append(["Priority", spec.get("priority", "") if spec else ""])
    ws.append(["Phase", getattr(feature, "phase", "")])

    # Confidence scores
    for art, label in [(spec, "Spec"), (plan, "Plan"), (tests, "Tests")]:
        if art:
            # Try to get from artifact object or dict
            score = art.get("_confidence", "") if isinstance(art, dict) else ""
            ws.append([f"Confidence ({label})", score])

    ws.append(["Total Turns", getattr(feature, "total_turns", 0)])
    ws.append(["Est. Hours Saved", f"{getattr(feature, 'estimated_hours_saved', 0):.1f}h"])
    if trace and trace.get("coverage_percent") is not None:
        ws.append(["Traceability Coverage", f"{trace['coverage_percent']}%"])
    ws.append(["Created", str(getattr(feature, "created_at", ""))])


def _build_stories(ws, spec, feature):
    headers = ["Issue Type", "Summary", "Description", "Priority", "Parent", "Story ID", "Acceptance Criteria", "Edge Cases"]
    ws.append(headers)
    _style_headers(ws, len(headers))

    fname = spec.get("feature_name", getattr(feature, "description", ""))
    ws.append(["Epic", fname, spec.get("business_context", ""), spec.get("priority", ""), "", "", "", ""])

    for story in spec.get("user_stories", []):
        summary = f"[{story.get('id', '')}] As {story.get('role', '')}, I want {story.get('action', '')}"
        desc = f"As a {story.get('role', '')}, I want {story.get('action', '')}, so that {story.get('benefit', '')}."
        acs = "; ".join(
            f"GIVEN {ac.get('given', '')} WHEN {ac.get('when', '')} THEN {ac.get('then', '')}"
            for ac in story.get("acceptance_criteria", []) if isinstance(ac, dict)
        )
        edges = "; ".join(spec.get("edge_cases", []))
        ws.append(["Story", summary[:250], desc, "High", fname, story.get("id", ""), acs, edges])


def _build_tasks(ws, plan):
    headers = ["Issue Type", "ID", "Title", "Description", "Parent Story", "Estimated Hours", "Affected Files", "Risks"]
    ws.append(headers)
    _style_headers(ws, len(headers))

    risks_str = "; ".join(
        f"[{r.get('severity', '').upper()}] {r.get('description', '')}"
        for r in plan.get("risks", []) if isinstance(r, dict)
    )

    for st in plan.get("subtasks", []):
        files = ", ".join(r.get("file", "") for r in plan.get("affected_routes", []) if isinstance(r, dict))
        ws.append(["Sub-task", st.get("id", ""), st.get("title", ""), st.get("description", ""),
                    st.get("story_id", ""), st.get("estimated_hours", 0), files, risks_str])


def _build_tests(ws, tests):
    headers = ["Suite", "Type", "ID", "Title", "Story", "Priority", "Preconditions", "Steps", "Expected Result", "Automated"]
    ws.append(headers)
    _style_headers(ws, len(headers))

    for suite in tests.get("test_suites", []):
        for tc in suite.get("test_cases", []):
            pre = "; ".join(tc.get("preconditions", [])) if isinstance(tc.get("preconditions"), list) else str(tc.get("preconditions", ""))
            steps = "; ".join(tc.get("steps", [])) if isinstance(tc.get("steps"), list) else str(tc.get("steps", ""))
            ws.append([
                suite.get("name", ""), suite.get("type", ""), tc.get("id", ""), tc.get("title", ""),
                suite.get("story_id", ""), tc.get("priority", ""), pre, steps,
                tc.get("expected_result", ""), "Yes" if tc.get("automated") else "No",
            ])


def _build_traceability(ws, trace):
    headers = ["AC ID", "Type", "Message", "Status"]
    ws.append(headers)
    _style_headers(ws, len(headers))

    # Covered ACs
    for story_id in trace.get("stories_in_spec", []):
        in_plan = story_id in trace.get("stories_in_plan", [])
        in_tests = story_id in trace.get("stories_in_tests", [])
        status = "Covered" if in_plan and in_tests else "Gap"
        ws.append([story_id, "", "", status])

    for gap in trace.get("gaps", []):
        ws.append([gap.get("ac_id", ""), gap.get("type", ""), gap.get("message", ""), "GAP"])

    ws.append([])
    ws.append(["Coverage", f"{trace.get('coverage_percent', 0)}%"])
    ws.append(["Total ACs", trace.get("total_acceptance_criteria", 0)])
    ws.append(["Covered", trace.get("covered", 0)])
    ws.append(["Gaps", len(trace.get("gaps", []))])


def _build_knowledge(ws, entries):
    headers = ["Type", "Title", "Content", "Tags", "Created"]
    ws.append(headers)
    _style_headers(ws, len(headers))

    for e in entries:
        tags = ", ".join(e.tags or []) if hasattr(e, "tags") and e.tags else ""
        ws.append([
            e.entry_type if hasattr(e, "entry_type") else e.get("entry_type", ""),
            e.title if hasattr(e, "title") else e.get("title", ""),
            (e.content if hasattr(e, "content") else e.get("content", ""))[:500],
            tags,
            str(e.created_at if hasattr(e, "created_at") else e.get("created_at", "")),
        ])


def _build_scaffold(ws, scaffold):
    headers = ["File Path", "Language", "Subtask", "Functions", "Description"]
    ws.append(headers)
    _style_headers(ws, len(headers))

    for f in scaffold.get("scaffold_files", []):
        ws.append([
            f.get("path", ""), f.get("language", ""), f.get("subtask_id", ""),
            ", ".join(f.get("functions", [])), f.get("description", ""),
        ])


# ── Markdown Export ──────────────────────────────────────────────────

def export_feature_markdown(feature, spec, plan, tests, scaffold,
                            knowledge_entries, traceability_report) -> str:
    lines = []
    fname = spec.get("feature_name", "") if spec else getattr(feature, "description", "")
    phase = getattr(feature, "phase", "")
    priority = spec.get("priority", "") if spec else ""

    lines.append(f"# Feature: {fname}")
    lines.append(f"> Priority: {priority} | Phase: {phase}")
    hours = getattr(feature, "estimated_hours_saved", 0)
    if hours:
        lines.append(f"> Est. Hours Saved: {hours:.1f}h | Turns: {getattr(feature, 'total_turns', 0)}")
    if traceability_report and traceability_report.get("coverage_percent") is not None:
        lines.append(f"> Traceability: {traceability_report['coverage_percent']}% coverage | {len(traceability_report.get('gaps', []))} gap(s)")
    lines.append("")
    lines.append("---")
    lines.append("")

    # Spec
    if spec:
        lines.append("## 1. Specification")
        lines.append("")
        if spec.get("business_context"):
            lines.append("### Business Context")
            lines.append(spec["business_context"])
            lines.append("")
        if spec.get("personas"):
            lines.append("### Personas")
            for p in spec["personas"]:
                if isinstance(p, dict):
                    lines.append(f"- **{p.get('name', '')}**: {p.get('description', '')}")
            lines.append("")
        if spec.get("user_stories"):
            lines.append("### User Stories")
            for s in spec["user_stories"]:
                lines.append(f"\n#### {s.get('id', '')}: As {s.get('role', '')}, I want {s.get('action', '')}")
                lines.append(f"**Benefit:** {s.get('benefit', '')}")
                lines.append("")
                if s.get("acceptance_criteria"):
                    lines.append("**Acceptance Criteria:**")
                    for i, ac in enumerate(s["acceptance_criteria"], 1):
                        if isinstance(ac, dict):
                            lines.append(f"{i}. GIVEN {ac.get('given', '')} WHEN {ac.get('when', '')} THEN {ac.get('then', '')}")
                    lines.append("")
        if spec.get("edge_cases"):
            lines.append("### Edge Cases")
            for e in spec["edge_cases"]:
                lines.append(f"- {e}")
            lines.append("")
        if spec.get("non_functional_requirements"):
            lines.append("### Non-Functional Requirements")
            for n in spec["non_functional_requirements"]:
                lines.append(f"- {n}")
            lines.append("")
        lines.append("---")
        lines.append("")

    # Plan
    if plan:
        lines.append("## 2. Technical Plan")
        lines.append("")
        if plan.get("affected_routes"):
            lines.append("### Affected Routes")
            for r in plan["affected_routes"]:
                if isinstance(r, dict):
                    lines.append(f"- `{r.get('method', '')} {r.get('path', '')}` in `{r.get('file', '')}` — {r.get('change', '')}")
            lines.append("")
        if plan.get("subtasks"):
            lines.append("### Subtasks")
            lines.append("| ID | Title | Story | Hours |")
            lines.append("|---|---|---|---|")
            for t in plan["subtasks"]:
                lines.append(f"| {t.get('id', '')} | {t.get('title', '')} | {t.get('story_id', '')} | {t.get('estimated_hours', '')}h |")
            lines.append("")
        if plan.get("risks"):
            lines.append("### Risks")
            for r in plan["risks"]:
                if isinstance(r, dict):
                    lines.append(f"- **[{r.get('severity', '').upper()}]** {r.get('description', '')} — {r.get('mitigation', '')}")
            lines.append("")
        lines.append("---")
        lines.append("")

    # Tests
    if tests and tests.get("test_suites"):
        lines.append("## 3. Test Plan")
        lines.append("")
        total = sum(len(s.get("test_cases", [])) for s in tests["test_suites"])
        lines.append(f"**{total} test cases** across {len(tests['test_suites'])} suites")
        lines.append("")
        for suite in tests["test_suites"]:
            lines.append(f"### {suite.get('name', '')} ({suite.get('type', '')})")
            for tc in suite.get("test_cases", []):
                lines.append(f"- **{tc.get('id', '')}: {tc.get('title', '')}** ({tc.get('priority', '')})")
            lines.append("")
        lines.append("---")
        lines.append("")

    # Traceability
    if traceability_report and traceability_report.get("status") == "complete":
        lines.append("## 4. Traceability Matrix")
        lines.append(f"Coverage: **{traceability_report.get('coverage_percent', 0)}%**")
        lines.append("")
        if traceability_report.get("gaps"):
            lines.append("### Gaps")
            for g in traceability_report["gaps"]:
                lines.append(f"- {g.get('message', '')}")
            lines.append("")
        lines.append("---")
        lines.append("")

    # Knowledge
    if knowledge_entries:
        lines.append("## 5. Knowledge & Decisions")
        lines.append("")
        for e in knowledge_entries:
            title = e.title if hasattr(e, "title") else e.get("title", "")
            etype = e.entry_type if hasattr(e, "entry_type") else e.get("entry_type", "")
            content = e.content if hasattr(e, "content") else e.get("content", "")
            lines.append(f"### {title}")
            lines.append(f"**Type:** {etype}")
            lines.append(content[:500])
            lines.append("")
        lines.append("---")
        lines.append("")

    lines.append(f"*Generated by Synapse AI Engineering Governance Platform*")
    lines.append(f"*Exported: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC*")

    return "\n".join(lines)
